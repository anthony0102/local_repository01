# -*-coding:utf-8-*-


from openpyxl import load_workbook


wb = load_workbook(r'data.xlsx')
ws = wb.active

print(ws['A1'], ws['B1'].value, ws['C1'].value)

for each_row in ws.rows:
    print(each_row[1].value, end=' ')

print('\n')

for each_column in ws.columns:
    print(each_column[0].value, end='|')
