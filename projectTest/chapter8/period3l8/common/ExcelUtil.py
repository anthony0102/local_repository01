# -*-coding:utf-8-*-
# 封装读取数据文件的方法
# util: utility 实用程序;公用程序(UTIL)
#       utilization 利用，使用
import os
from openpyxl import load_workbook


class ExcelUtil(object):
    def __init__(self, excel_path=None, sheet_name=None):
        """获取excel工作表"""

        # 如果默认没有传入excel_path参数，则在当前目录上一级的data中找目标xlsx
        if excel_path == None:
            current_path = os.path.abspath(os.path.dirname(__file__))
            self.excel_path = current_path + '/../data/1.xlsx'
            # self.excel_path = current_path + '/../data/casedata.xlsx'
            # !!! pycharm中新建的casedata.xlsx文件不能被成功读取，在WPS中另存为1.xlsx则读取成功

        # 如果传入了excel_path就把值传给自己的excel_path
        else:
            self.excel_path = excel_path

        # sheet_name 是xlsx表格中单个工作表的名称，没改的话默认是Sheet1
        # 如果没有传入值，没有更改，就是默认值
        if sheet_name == None:
            self.sheet_name = "Sheet1"
        else:
            self.sheet_name = sheet_name

        # 打开工作表
        self.data = load_workbook(self.excel_path)
        self.sheet = self.data[self.sheet_name]

    def get_data(self):
        """
        获取文件数据
        每一行数据一个list，所有的数据一个大list
        """

        rows = self.sheet.rows
        # 表中数据占据的行数
        row_num = self.sheet.max_row
        # 表中数据占据的列数
        col_num = self.sheet.max_column

        if row_num <= 1:
            print("总行数小于1，没有数据")
        else:
            # print(row_num)  # 13
            # print(col_num)  # 6
            case_all = []
            for row in rows:
                case = []
                for i in range(col_num):
                    case.append(row[i].value)
                case_all.append(case)
            return case_all


if __name__ == '__main__':
    sheet = 'Login'
    file = ExcelUtil(sheet_name=sheet)
    print(file.get_data())
