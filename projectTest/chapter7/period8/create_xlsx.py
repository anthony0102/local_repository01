# -*-coding:utf-8-*-
# pip install openpyxl


import datetime
from openpyxl import Workbook


# 创建文件对象，也就是创建xlsx文件
wb = Workbook()
# 获取默认工作表sheet
ws = wb.active

ws['A1'] = "编号"
ws['B1'] = "任务"
ws['C1'] = "创建时间"

ws.append({'A': 10001, 'B': "测试任务", 'C': datetime.datetime.now()})

wb.save('data.xlsx')
